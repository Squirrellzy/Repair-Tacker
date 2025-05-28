
import streamlit as st
import pandas as pd
from datetime import datetime
import os
from github import Github
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Constants
GITHUB_REPO = "Squirrellzy/Repair-Tacker"
GITHUB_TOKEN = st.secrets["GITHUB_TOKEN"]
USERS = {
    "aci": "mars",
    "usps": "mars",
    "retiina": "mars",
    "admin": "Ret1B"
}

# Login
st.title("Login")
login_success = False
username_input = st.text_input("Username", key="login_user")
password_input = st.text_input("Password", type="password", key="login_pass")
site = st.selectbox("Select Site", ["Indy", "Chicago", "Atlanta"], key="site_selector")
EXCEL_FILE_TEMPLATE = "cc_comments_log_{site}.xlsx"

if username_input in USERS and password_input == USERS[username_input]:
    login_success = True
    logged_user = username_input
    EXCEL_FILE = EXCEL_FILE_TEMPLATE.format(site=site)
    st.success("Login successful!")

if login_success:
    def format_excel_file(path):
        wb = load_workbook(path)
        ws = wb.active
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 2
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName="CCLogTable", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)
        formatted_path = f"cc_comments_log_formatted_{site}.xlsx"
        wb.save(formatted_path)
        return formatted_path

    if logged_user == "admin":
        st.title("Admin Panel - Full Log Viewer")
        selected_admin_site = st.selectbox("View logs for site:", ["Indy", "Chicago", "Atlanta"], key="admin_site_selector")
        admin_excel_file = EXCEL_FILE_TEMPLATE.format(site=selected_admin_site)

        if os.path.exists(admin_excel_file):
            df_admin = pd.read_excel(admin_excel_file)
            st.dataframe(df_admin)

            formatted = format_excel_file(admin_excel_file)
            with open(formatted, "rb") as f:
                st.download_button(
                    label="Download Full Excel Log",
                    data=f,
                    file_name=formatted,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        st.stop()

    if not os.path.exists(EXCEL_FILE):
        df_init = pd.DataFrame(columns=["Date", "User", "CC_Subsection", "Description"])
        df_init.to_excel(EXCEL_FILE, index=False)

    st.title("Collection Conveyor Comment Logger")
    cc_number = st.selectbox("Select Collection Conveyor", [f"CC-{i}" for i in range(1, 78)], key="main_cc_selector")
    comment_1 = st.text_area("A side 1 Comment", key="main_comment_1")
    comment_2 = st.text_area("2 Comment", key="main_comment_2")
    comment_3 = st.text_area("3 Comment", key="main_comment_3")
    comment_4 = st.text_area("B side 4 Comment", key="main_comment_4")

    if st.button("Submit Comment", key="submit_btn"):
        entries = []
        date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if comment_1.strip():
            entries.append([date, logged_user, f"{cc_number}-A side 1", comment_1.strip()])
        if comment_2.strip():
            entries.append([date, logged_user, f"{cc_number}-2", comment_2.strip()])
        if comment_3.strip():
            entries.append([date, logged_user, f"{cc_number}-3", comment_3.strip()])
        if comment_4.strip():
            entries.append([date, logged_user, f"{cc_number}-B side 4", comment_4.strip()])

        if entries:
            new_df = pd.DataFrame(entries, columns=["Date", "User", "CC_Subsection", "Description"])
            df_existing = pd.read_excel(EXCEL_FILE)
            df_combined = pd.concat([df_existing, new_df], ignore_index=True)
            df_combined.to_excel(EXCEL_FILE, index=False)
            st.success("Comment(s) logged successfully!")

            try:
                g = Github(GITHUB_TOKEN)
                repo = g.get_repo(GITHUB_REPO)
                contents = None
                try:
                    contents = repo.get_contents(EXCEL_FILE)
                except:
                    contents = None

                with open(EXCEL_FILE, "rb") as f:
                    content = f.read()

                if contents:
                    repo.update_file(EXCEL_FILE, "Update log", content, contents.sha)
                else:
                    repo.create_file(EXCEL_FILE, "Create log", content)
                st.success("Excel file updated on GitHub!")
            except Exception as e:
                st.warning(f"Failed to push to GitHub: {e}")
        else:
            st.info("No comments entered.")

    st.markdown("---")
    st.header("Download Log")
    if os.path.exists(EXCEL_FILE):
        formatted = format_excel_file(EXCEL_FILE)
        with open(formatted, "rb") as f:
            st.download_button(
                label="Download Excel Log",
                data=f,
                file_name=formatted,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
else:
    st.warning("Please enter valid credentials.")
